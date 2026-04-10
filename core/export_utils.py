"""
export_utils.py — CSV / Excel import-export for attendance data
"""

import pandas as pd
import csv
import os
from datetime import datetime
from pathlib import Path

EXPORT_DIR = Path(__file__).parent.parent / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


# ── Export ────────────────────────────────────────────────────────────────────

def export_attendance_csv(records: list[dict], session_name: str = "") -> str:
    """Export attendance records to CSV. Returns file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = session_name.replace(" ", "_") if session_name else "session"
    path = EXPORT_DIR / f"attendance_{safe_name}_{ts}.csv"

    if not records:
        # Create empty file with headers
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["name","employee_id","department","check_in","confidence","liveness_score","status"])
            writer.writeheader()
        return str(path)

    df = pd.DataFrame(records)
    cols = ["name","employee_id","department","check_in","confidence","liveness_score","status"]
    existing = [c for c in cols if c in df.columns]
    df = df[existing]
    df.to_csv(path, index=False)
    return str(path)


def export_attendance_excel(records: list[dict], session_name: str = "") -> str:
    """Export attendance records to Excel with formatting. Returns file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = session_name.replace(" ", "_") if session_name else "session"
    path = EXPORT_DIR / f"attendance_{safe_name}_{ts}.xlsx"

    df = pd.DataFrame(records) if records else pd.DataFrame(
        columns=["name","employee_id","department","check_in","confidence","liveness_score","status"]
    )
    cols = ["name","employee_id","department","check_in","confidence","liveness_score","status"]
    existing = [c for c in cols if c in df.columns]
    df = df[existing]

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
        ws = writer.sheets["Attendance"]

        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    return str(path)


def export_persons_csv(persons: list[dict]) -> str:
    """Export registered persons to CSV."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = EXPORT_DIR / f"persons_{ts}.csv"
    df = pd.DataFrame(persons)
    cols = ["id","name","employee_id","department","email","created_at","is_active"]
    existing = [c for c in cols if c in df.columns]
    df[existing].to_csv(path, index=False)
    return str(path)


def export_all_sessions_excel(sessions: list[dict], all_attendance: list[dict]) -> str:
    """Export all sessions with their attendance to a multi-sheet Excel."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = EXPORT_DIR / f"full_report_{ts}.xlsx"

    from openpyxl.styles import Font, PatternFill, Alignment
    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        # Sheet 1: Sessions summary
        df_s = pd.DataFrame(sessions) if sessions else pd.DataFrame(
            columns=["id","name","date","start_time","end_time"]
        )
        df_s.to_excel(writer, index=False, sheet_name="Sessions")

        # Sheet 2: All Attendance
        df_a = pd.DataFrame(all_attendance) if all_attendance else pd.DataFrame(
            columns=["name","employee_id","department","session_name","check_in","confidence","status"]
        )
        cols = ["name","employee_id","department","session_name","check_in","confidence","liveness_score","status"]
        existing = [c for c in cols if c in df_a.columns]
        df_a[existing].to_excel(writer, index=False, sheet_name="All Attendance")

        # Style
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ml = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = ml + 4

    return str(path)


# ── Import ────────────────────────────────────────────────────────────────────

def import_persons_from_csv(filepath: str) -> list[dict]:
    """
    Import persons from CSV.
    Expected columns: name, employee_id, department (optional), email (optional)
    Returns list of person dicts ready to insert.
    """
    df = pd.read_csv(filepath)
    df.columns = [c.strip().lower() for c in df.columns]
    required = {"name", "employee_id"}
    if not required.issubset(set(df.columns)):
        raise ValueError(f"CSV must have columns: {required}. Found: {set(df.columns)}")
    df = df.fillna("")
    return df.to_dict("records")


def list_exports() -> list[dict]:
    """List all exported files."""
    files = []
    for f in sorted(EXPORT_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.suffix in {".csv", ".xlsx"}:
            files.append({
                "name": f.name,
                "path": str(f),
                "size_kb": round(f.stat().st_size / 1024, 1),
                "modified": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return files
